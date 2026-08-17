"""Safe importer for the existing CVE research workbook (Step 11).

The workbook (telecom_vpn_cve_zero_click.xlsx, " CVE Intelligence" sheet - note
the leading space, the sheet's real name) is no longer the runtime database,
but remains valuable research evidence: its manually-researched "Zero-Click
Relevance" column is exactly the one signal zeroshield.intelligence.priority's
domain-relevance scoring cannot derive from NVD/KEV/EPSS automatically.

Import produces the SAME NormalisedContribution shape every connector
produces (source=MANUAL_IMPORT), merged through the identical
zeroshield.intelligence.dedup.merge()/repository path - never a special case.
This module NEVER creates, modifies, or triggers a run of an
ExperimentDefinition: it has no import of zeroshield.runners/orchestration/
worker, and produces only Vulnerability/ValidationCandidate data, matching
Step 11's "Never automatically execute experiments from imported rows."

Requires the optional "excel" extra (openpyxl) - imported lazily inside
import_workbook so it stays optional for every other code path.
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from zeroshield.intelligence.normalisation import NormalisedContribution
from zeroshield.models.enums import ZeroClickRelevance
from zeroshield.models.vulnerability import VulnerabilitySourceName

_CVE_ID_RE = re.compile(r"^CVE-\d{4}-\d{4,7}$")
_CWE_RE = re.compile(r"CWE-\d+")
_DEFAULT_SHEET_NAME = " CVE Intelligence"  # the workbook's actual sheet name, leading space included

_ZERO_CLICK_MAP = {"high": ZeroClickRelevance.HIGH, "medium": ZeroClickRelevance.MEDIUM, "low": ZeroClickRelevance.LOW}


class ExcelImportError(Exception):
    pass


@dataclass(frozen=True)
class SkippedRow:
    row_number: int
    reason: str


@dataclass(frozen=True)
class ImportResult:
    contributions: list[NormalisedContribution]
    skipped: list[SkippedRow]


def _clean_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_cvss(value: Any) -> float | None:
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    return score if 0.0 <= score <= 10.0 else None


def _parse_percentage(value: Any) -> float | None:
    """Handles both "99.57%" (string, as in the workbook) and a bare 0-1 float."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value) if 0.0 <= float(value) <= 1.0 else None
    text = str(value).strip().rstrip("%")
    try:
        parsed = float(text)
    except ValueError:
        return None
    parsed = parsed / 100.0 if parsed > 1.0 else parsed
    return parsed if 0.0 <= parsed <= 1.0 else None


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("yes", "true", "y"):
        return True
    if text in ("no", "false", "n"):
        return False
    return None


def import_workbook(path: Path, *, sheet_name: str = _DEFAULT_SHEET_NAME) -> ImportResult:
    """Reads every data row of the given sheet, producing one
    NormalisedContribution per row with a recognisable CVE ID (source=
    MANUAL_IMPORT) - rows with no/malformed CVE ID are reported in `skipped`,
    never silently dropped (Step 12: "test... duplicates, partial data")."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelImportError(
            "openpyxl is required to import the CVE research workbook - install the 'excel' extra"
        ) from exc

    if not path.is_file():
        raise ExcelImportError(f"workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ExcelImportError(
            f"sheet '{sheet_name}' not found in {path}; available sheets: {workbook.sheetnames}"
        )
    sheet = workbook[sheet_name]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ExcelImportError(f"sheet '{sheet_name}' is empty") from None

    columns = {str(name).strip(): idx for idx, name in enumerate(header) if name}
    if "CVE ID" not in columns:
        raise ExcelImportError(f"required column 'CVE ID' not found in sheet '{sheet_name}' header")

    def _col(row: tuple[Any, ...], name: str) -> Any:
        idx = columns.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    contributions: list[NormalisedContribution] = []
    skipped: list[SkippedRow] = []

    for row_number, row in enumerate(rows_iter, start=2):  # row 1 is the header
        if row is None or all(v is None for v in row):
            continue

        cve_id = _clean_str(_col(row, "CVE ID"))
        if not cve_id or not _CVE_ID_RE.match(cve_id):
            skipped.append(SkippedRow(row_number=row_number, reason=f"missing/malformed CVE ID: {cve_id!r}"))
            continue

        references = [
            url for url in (_clean_str(_col(row, f"Source URL {n}")) for n in (1, 2, 3)) if url
        ]
        cwe_text = _clean_str(_col(row, "CWE")) or ""
        cwe_ids = _CWE_RE.findall(cwe_text)

        zero_click_raw = _clean_str(_col(row, "Zero-Click Relevance"))
        zero_click = _ZERO_CLICK_MAP.get(zero_click_raw.lower()) if zero_click_raw else None

        vendor = _clean_str(_col(row, "Vendor"))
        product = _clean_str(_col(row, "Product / Component"))
        products: list[tuple[str, str, str | None]] = [(vendor, product, None)] if vendor and product else []

        kev = _parse_bool(_col(row, "CISA KEV?"))

        contributions.append(
            NormalisedContribution(
                cve_id=cve_id,
                source=VulnerabilitySourceName.MANUAL_IMPORT,
                source_identifier=None,
                description=_clean_str(_col(row, "Trust Boundary Affected")),
                published_at=None,
                last_modified_at=None,
                cvss_score=_parse_cvss(_col(row, "CVSS Score")),
                cvss_vector=None,
                cvss_version=None,
                epss_score=_parse_percentage(_col(row, "EPSS Score")),
                epss_percentile=None,
                epss_date=None,
                kev_listed=kev,
                kev_date_added=None,
                kev_due_date=None,
                kev_known_ransomware=None,
                cwe_ids=cwe_ids,
                vendor=vendor,
                products=products,
                references=references,
                zero_click_relevance=zero_click,
            )
        )

    return ImportResult(contributions=contributions, skipped=skipped)


@dataclass(frozen=True)
class ImportSummary:
    fetched_count: int
    created_count: int
    updated_count: int
    unchanged_count: int
    failed_count: int
    skipped: list[SkippedRow]


def import_and_merge(
    path: Path,
    *,
    repository: Any,
    sheet_name: str = _DEFAULT_SHEET_NAME,
) -> ImportSummary:
    """Reads the workbook and merges every row into the same Vulnerability
    system of record every connector writes to (source=MANUAL_IMPORT) -
    never touches ExperimentDefinition/experiments/ in any way (Step 11).
    `repository` is a zeroshield.intelligence.repository.VulnerabilityRepository,
    typed as Any here to keep this module importable without sqlalchemy
    installed unless import_and_merge is actually called.
    """
    from zeroshield.intelligence.dedup import merge

    result = import_workbook(path, sheet_name=sheet_name)
    created = updated = unchanged = 0
    for contribution in result.contributions:
        existing = repository.get_vulnerability(contribution.cve_id)
        merged = merge(existing, contribution)
        repository.upsert_vulnerability(merged.vulnerability)
        repository.upsert_source_record(merged.source_record)
        repository.append_history(merged.history)
        if merged.products:
            repository.upsert_products(contribution.cve_id, merged.products)
        if merged.is_new:
            created += 1
        elif merged.history:
            updated += 1
        else:
            unchanged += 1

    return ImportSummary(
        fetched_count=len(result.contributions),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        failed_count=len(result.skipped),
        skipped=result.skipped,
    )
